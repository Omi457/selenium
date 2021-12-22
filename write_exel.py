import openpyxl

path = "C://Users//Admin//Desktop//writeexcel.xlsx"

workbook = openpyxl.load_workbook(path)
# sheet = workbook.active
# for r in range(1, 6):
#     for c in range(1 , 5):
#         sheet.cell(row = r, column= c).value=("data1")

# workbook.save(path)



sheet = workbook.active
sheet = workbook.create_sheet("add_data" ,1 )  ## used for creat the sheet 
data = [("SR.NO","Name","Age"),(1,"omkar", 23),(2, "amol", 24),(3, "vikas", 22)]
for i in data:
    sheet.append(i)

workbook.save(path)