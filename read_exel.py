from selenium import webdriver 
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
import openpyxl
        
path = "C://Users//Admin//Desktop//TargetsShrinkingIsland.xlsx"

workbook = openpyxl.load_workbook(path)

# sheet = workbook.active
sheet = workbook.get_sheet_by_name("Sheet1")

row = sheet.max_row      ##9
clos = sheet.max_column  ##4

# print(row)
# print(clos)


for r in range(1, row+1):
    for c in range(1, clos+1):
        print(sheet.cell(row = r, column = c). value, end = "  ")

    print()
    